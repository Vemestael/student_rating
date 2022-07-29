from datetime import datetime
from uuid import uuid4

import django.contrib.auth as da
from django.conf import settings
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.core.handlers.wsgi import WSGIRequest
from django.core.paginator import Paginator
from django.core.serializers import serialize
from django.db.models import QuerySet
from django.http import JsonResponse
from django.shortcuts import render, redirect
from django.views.generic.base import TemplateView, View
from openpyxl import load_workbook
from rest_framework.response import Response
from rest_framework.views import APIView

import main.forms as forms
import main.models as model
from main.serializers import ExtraPointSerializer


def is_ajax(request: WSGIRequest) -> bool:
    """
    Returns True if the request was made via an XMLHttpRequest
    """
    return request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest'


def get_semester_date_range(session_time: str, year: str) -> list[str]:
    """
    Return a range of semester dates depending on the selected exam session (summer or winter)
    """
    if 'summer' in session_time:
        start = year + "-03-01"
        end = year + "-08-01"
        return [start, end]
    elif 'winter' in session_time:
        start = year + "-08-02"
        end = str(int(year) + 1) + "-02-28"
        return [start, end]


def get_rating(faculty: int, session: str, year: str) -> QuerySet:
    """
    Returns the rating of students in a particular faculty for the selected session
    """
    date = get_semester_date_range(session, year)
    rating = model.Rating.objects.filter(date__range=date, faculty=faculty).values('id', 'full_name', 'group',
                                                                                   'session', 'extra', 'total')
    return rating.order_by('-total')


def read_workbook(file_name: str, action: str) -> None:
    """
    Reads student rating data from an Excel file and adds or updates records to the database
    """
    file_dir = str(settings.MEDIA_ROOT) + "/media/excel/" + file_name
    workbook = load_workbook(file_dir)
    worksheet = workbook[workbook.sheetnames[0]]
    faculties = {
        "ФИТ": 1,
        "ФІТ": 1,
        "ЕкФ": 2,
        "ЕнФ": 3,
        "МФ": 4,
        "СГФ": 5,
        "ФМЗ": 6,
        "ФТТ": 7,
        "ФИМП": 8,
        "ФІМП": 8,
    }
    faculty = faculties.get(worksheet.cell(row=2, column=2).value)
    faculty = model.Faculty.objects.get(id=faculty)

    if action == 'new':
        for row in worksheet.iter_rows(min_row=2):
            temp = []
            for cells in row:
                temp.append(cells.value)
            table_entry = model.Rating()
            table_entry.full_name = temp[0]
            table_entry.faculty = faculty
            table_entry.group = temp[2]
            table_entry.session = temp[3]
            table_entry.extra = temp[4]
            table_entry.total = float(temp[3]) + float(temp[4])
            table_entry.save()
    if action == 'update':
        for row in worksheet.iter_rows(min_row=2):
            temp = []
            for cells in row:
                temp.append(cells.value)
            table_entry = model.Rating.objects.get(faculty=faculty, group=temp[2], full_name=temp[0])
            table_entry.session = temp[3]
            table_entry.extra = temp[4]
            table_entry.total = float(temp[3]) + float(temp[4])
            table_entry.save()


class Index(TemplateView):
    """
    View to handle home page requests
    The main page displays the university's student rating and filtering tools for output
    """
    template_name = 'main/index.html'

    def get_context_data(self, **kwargs):
        """
        Checks the validity of the data from the filters form and returns a filtered students rating
        """
        form = forms.FilterForm(self.request.GET)
        if form.is_valid():
            faculty = form.cleaned_data.get('faculty')
            session = form.cleaned_data.get('session')
            year = form.cleaned_data.get('year_picker')
            records_per_page = form.cleaned_data.get('records_per_page')
            rating = get_rating(faculty, session, year)
        else:
            form = forms.FilterForm(initial={'year_picker': str(datetime.now().year)})
            faculty = form['faculty'].initial
            session = form['session'].initial
            year = form['year_picker'].initial
            records_per_page = form['records_per_page'].initial
            rating = get_rating(faculty, session, year)

        paginator = Paginator(rating, records_per_page)
        page_number = self.request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        context = {"page_obj": page_obj, "form": form}
        return context


class UploadCertificate(View):
    """
    View to handle upload-certificate request
    The UploadCertificate request allows the student to upload a certificate of participation in university events
    to increase the number of points in the ranking
    """

    def post(self, request):
        """
        Handles the POST request with the uploaded file, uploading information about it to the database
        """
        content_type = ["image/jpeg", "image/jpg", "image/png"]
        file = request.FILES['upload_file']
        if file.content_type in content_type:
            student = model.Rating.objects.get(pk=request.POST['student_id'])
            table_entry = model.Certificate(uploaded_by_student=student, certificate_file=file)
            table_entry.save()
        return redirect('/')


class GetDetails(APIView):
    """
    View to handle an ajax request to provide additional information about the student's rating
    The GetDetails request gives detailed information about the student's additional points in the rating
    """

    def get(self, request):
        """
        Handles an AJAX request, gives detailed information about the student's additional points in the rating
        """
        if is_ajax(request):
            student_id = request.GET['student']
            details = model.ExtraPoint.objects.filter(student_id=student_id)
            details = ExtraPointSerializer(instance=details, many=True)
            return Response(details.data)


class Login(View):
    """
    View to handle login page
    """

    def get(self, request):
        """
        Handles a GET request, gives a page for authorization
        """
        context = {'error': ''}
        return render(request, 'main/sign_in.html', context)

    def post(self, request):
        """
        Handles a POST request, checking the validity of user data for authorization
        """
        username = request.POST['login']
        password = request.POST['password']
        user = da.authenticate(request, username=username, password=password)
        if user is not None:
            da.login(request, user)
            return redirect('/')
        else:
            context = {'error': 'invalid login'}
            return render(request, 'main/sign_in.html', context)


class SignUp(View):
    """
    View to handle sign-up page
    """

    def get(self, request):
        """
        Handles a GET request, gives a page for registration
        """
        da.logout(request)
        invite_key_status = ""
        error = ""
        context = {"invite_key": invite_key_status, "error": error}
        return render(request, 'main/sign_up.html', context)

    def post(self, request):
        """
        Handles a POST request, checking the validity of user data for registration
        """
        da.logout(request)
        invite_key_status = ""
        error = ""
        if request.POST['invite_key_status'] != 'OK':
            invite_key = request.POST['invite_key']
            if model.InviteKey.objects.filter(invite_key__exact=invite_key).count():
                invite_key_status = "OK"
            else:
                error = 'invalid key'
        else:
            full_name = request.POST['full_name'].split(' ')
            first_name = full_name[0]
            last_name = ''
            if len(full_name) > 1:
                last_name = full_name[1]
            if len(first_name) == 0 or len(last_name) == 0:
                error = 'enter your full name'
                context = {"invite_key": "OK", "error": error}
                return render(request, 'main/sign_up.html', context)

            faculty = request.POST.get('faculty')
            if faculty is None:
                error = 'enter your faculty'
                context = {"invite_key": "OK", "error": error}
                return render(request, 'main/sign_up.html', context)

            email = request.POST['email']
            login = request.POST['login']
            password = request.POST['password']

            user = User.objects.create_user(login, email, password, first_name=first_name, last_name=last_name)
            user.save()
            return redirect('/')
        context = {"invite_key": invite_key_status, "error": error}
        return render(request, 'main/sign_up.html', context)


class InviteKeyGen(View):
    """
    View to handle a request to generate a new invite key
    The InviteKeyGen request allows the administrator to generate new keys for new users to register
    """

    def post(self, request):
        """
        Handles a POST request, generating new invite keys for the administrator to register new users
        """
        if request.user.is_authenticated:
            if request.user.is_superuser:
                invite_key = ""
                if request.method == "POST":
                    invite_key = uuid4()
                    a_record = model.InviteKey(invite_key=invite_key)
                    a_record.save()
                context = {"invite_key": invite_key}
                return render(request, 'admin/invite_key_gen.html', context)
        return redirect('/')


class AddRating(View):
    """
    View to handle add-rating page
    The AddRating page allows authorized users to add manually or from an Excel file new records with student ratings
    """

    def get(self, request):
        """
        Handles a GET request, gives a page for add manually or from an Excel file
        a new records with student ratings
        """
        username = request.user.username
        form = forms.AddRatingForm()
        context = {'username': username, "form": form}
        return render(request, 'main/add_rating.html', context)

    def post(self, request):
        """
        Handles a POST request, adding new records with student ratings
        """
        username = request.user.username
        content_type = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "application/vnd.ms-excel"]
        if 'upload_rating' in request.POST:
            file = request.FILES['upload_file']
            if file.content_type in content_type:
                table_entry = model.ExcelFile(uploaded_by_user=username, excel_file=file)
                table_entry.save()
                read_workbook(file.name, 'new')
            return redirect('/add-rating')
        if 'add_rating' in request.POST:
            form = forms.AddRatingForm(request.POST)
            if form.is_valid():
                rating = form.save(commit=False)
                rating.total = form.cleaned_data.get('session') + form.cleaned_data.get('extra')
                rating.save()
                return redirect('/')
            else:
                context = {'username': username, "form": form}
                return render(request, 'main/add_rating.html', context)


class CheckCertificate(View):
    """
    View to handle check-certificate page
    The CheckCertificate page allows authorized users to check the certificates uploaded by the students
    for increase their sum of points in the rating
    """

    def get(self, request):
        """
        Handles a GET request, gives a page for check certificate
        """
        certificates = model.Certificate.objects.all()
        context = {'certificates': certificates}
        return render(request, 'main/check-certificate.html', context)

    def post(self, request):
        """
        Handles a POST request, rejecting the certificate or adding extra points to the student's rating
        and certificate to the advanced rating information
        """
        if is_ajax(request):
            record_id = request.POST['record_id']
            if request.POST['action'] == 'add':
                certificate = model.Certificate.objects.get(pk=record_id)
                student_id = certificate.uploaded_by_student
                point = request.POST['added_points']
                description = request.POST['activity']
                certificate_file = certificate.certificate_file
                student_rating = model.Rating.objects.get(pk=student_id.id)
                extra_points = model.ExtraPoint(student_id=student_rating, point=point, description=description,
                                                certificate=certificate_file)
                student_rating.extra += int(point)
                student_rating.total += int(point)
                student_rating.save()
                extra_points.save()
                certificate.delete()
            if request.POST['action'] == 'reject':
                model.Certificate.objects.get(pk=record_id).delete()

        certificates = model.Certificate.objects.all()
        context = {'certificates': certificates}
        return render(request, 'main/check-certificate.html', context)


class ChangeRating(View):
    """
    View to handle change-rating page
    The ChangeRating page allows authorized users to manually change students' ratings
    """

    def get(self, request):
        """
        Handles a GET request, gives information about the rating of the selected student
        """
        faculties = model.Faculty.objects.all()
        student = model.Rating.objects.filter(faculty=1)
        if is_ajax(request):
            if 'faculty' in request.GET:
                student = model.Rating.objects.filter(faculty=int(request.GET['faculty']))
            if 'student' in request.GET:
                student = model.Rating.objects.filter(pk=int(request.GET['student']))
            student = serialize('json', student)
            return JsonResponse(student, safe=False)
        context = {'faculties': faculties, 'students': student}
        return render(request, 'main/change_rating.html', context)

    def post(self, request):
        """
        Handles a POST request, changing the rating information of the selected student
        """
        faculties = model.Faculty.objects.all()
        student = model.Rating.objects.filter(faculty=1)
        student_id = int(request.POST['student_id'])
        session = float(request.POST['session'])
        extra_point = float(request.POST['added_points'])
        description = request.POST['activity']
        file = ''
        if 'upload_file' in request.FILES:
            file = request.FILES['upload_file']
        student_rating = model.Rating.objects.get(pk=student_id)
        extra_points = model.ExtraPoint(student_id=student_rating, point=extra_point, description=description,
                                        certificate=file)
        student_rating.session = session
        student_rating.extra += extra_point
        student_rating.total += extra_point
        student_rating.save()
        extra_points.save()

        context = {'faculties': faculties, 'students': student}
        return render(request, 'main/change_rating.html', context)


class ChangeFromFile(View):
    """
    View to handle change-from-file page
    The ChangeRating page allows authorized users to change student ratings
    by uploading an Excel file with the new rating
    """

    def post(self, request):
        """
        Handles a POST request, changing student rating information by reading the new data from the uploaded Excel file
        """
        username = request.user.username
        content_type = ["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        "application/vnd.ms-excel"]
        file = request.FILES['upload_file']
        if file.content_type in content_type:
            table_entry = model.ExcelFile(uploaded_by_user=username, excel_file=file)
            table_entry.save()
            read_workbook(file.name, 'update')
        return redirect('home')


class Profile(View):
    """
    View to handle profile page
    The Profile page allows authorized users to change personal information,
    and admins can additionally generate new invite keys.
    """

    def get(self, request):
        """
        Handles a GET request, gives a profile page
        """
        success = ''
        password_change_form = PasswordChangeForm(user=request.user)
        context = {"password_change_form": password_change_form, "success": success}
        return render(request, 'main/profile.html', context)

    def post(self, request):
        """
        Handles a POST request, updating user data and generating invitation keys for the administrator
        """
        success = ''
        context = {}
        if is_ajax(request):
            if request.user.is_superuser:
                invite_key = uuid4()
                a_record = model.InviteKey(invite_key=invite_key)
                a_record.save()
                context = {"invite_key": invite_key}
            return JsonResponse(context)
        else:
            user = User.objects.get(id=request.user.id)
            if 'update_profile' in request.POST:
                user.username = request.POST['username']
                user.first_name = request.POST['first_name']
                user.last_name = request.POST['last_name']
                user.save()
                success = 'profile updated successfully'
            if 'update_email' in request.POST:
                user.email = request.POST['email']
                user.save()
                success = 'email updated successfully'
            if 'update_password' in request.POST:
                form = PasswordChangeForm(user=request.user, data=request.POST or None)
                if form.is_valid():
                    form.save()
                    update_session_auth_hash(request, form.user)
                success = 'password updated successfully'
            context = {"success": success}
            return render(request, 'main/profile.html', context)
